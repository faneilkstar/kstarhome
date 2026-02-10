from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def style_header(ws, row=1):
    """Applique le style à l'en-tête"""
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

def auto_adjust_columns(ws):
    """Ajuste automatiquement la largeur des colonnes"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def export_etudiants_excel(etudiants, filename=None):
    """Exporte la liste des étudiants en Excel"""
    # Chemin relatif sécurisé par défaut
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # On s'assure que le dossier documents existe
        base_dir = os.path.abspath(os.path.dirname(__file__))
        # Remonter de 'app/utils' vers la racine pour trouver 'documents'
        # root/app/utils -> root
        root_dir = os.path.dirname(os.path.dirname(base_dir))
        doc_dir = os.path.join(root_dir, 'documents')

        if not os.path.exists(doc_dir):
            os.makedirs(doc_dir)

        filename = os.path.join(doc_dir, f"export_etudiants_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Étudiants"

    # En-têtes
    headers = ['Matricule', 'Nom', 'Prénom', 'Date Naissance', 'Classe', 'Statut', 'Date Inscription']
    ws.append(headers)
    style_header(ws)

    # Données
    for etudiant in etudiants:
        row = [
            etudiant.get_matricule() if hasattr(etudiant, 'get_matricule') else f"ETU{etudiant.id}",
            etudiant.nom,
            etudiant.prenom,
            etudiant.date_naissance.strftime('%d/%m/%Y') if etudiant.date_naissance else '-',
            etudiant.classe.nom_classe if etudiant.classe else '-',
            etudiant.statut_inscription,
            etudiant.date_inscription.strftime('%d/%m/%Y') if etudiant.date_inscription else '-'
        ]
        ws.append(row)

    auto_adjust_columns(ws)
    wb.save(filename)
    return filename

def export_statistiques_excel(stats_data, filename=None):
    """Exporte les statistiques globales en Excel"""
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_dir = os.path.abspath(os.path.dirname(__file__))
        root_dir = os.path.dirname(os.path.dirname(base_dir))
        doc_dir = os.path.join(root_dir, 'documents')

        if not os.path.exists(doc_dir):
            os.makedirs(doc_dir)

        filename = os.path.join(doc_dir, f"export_statistiques_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Vue d'ensemble"

    ws['A1'] = "STATISTIQUES GLOBALES"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')

    ws.append([])
    ws.append(['Indicateur', 'Valeur'])
    style_header(ws, row=3)

    ws.append(['Total Étudiants', stats_data.get('total_etudiants', 0)])
    ws.append(['Total Enseignants', stats_data.get('total_enseignants', 0)])
    ws.append(['Total UE', stats_data.get('total_ues', 0)])
    ws.append(['Total Classes', stats_data.get('total_classes', 0)])

    auto_adjust_columns(ws)
    wb.save(filename)
    return filename